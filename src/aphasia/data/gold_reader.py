"""读取 gold xlsx 三个关键工作表（原则 I：金标准保真，实时取数）。

- 打分汇总：每病例金标准与用途
- 信息量和流畅度：评分标准与 prompt 模板（key-value 布局）
- fakeasr：AI 合成对话及其目标分与理由
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

import openpyxl

from ..config import GOLD_XLSX

SHEET_SUMMARY = "打分汇总"
SHEET_CRITERION = "信息量和流畅度"
SHEET_FAKE = "fakeasr"

# 评分标准页必须存在的 key（缺失即中止——原则 I）
REQUIRED_CRITERION_KEYS = (
    "system",
    "field",
    "dialogue_scope",
    "dialogue_content",
    "diag_info_criterion",
    "diag_flue_criterion",
    "question_surfix",
)


@dataclass(frozen=True)
class GoldCase:
    """打分汇总页的一行病例金标准。"""

    case_id: str
    info_gold: int
    flue_gold: int
    usage: str
    language: str | None


@dataclass(frozen=True)
class FakeCase:
    """fakeasr 页的一条合成对话。"""

    info_gold: int
    flue_gold: int
    gen_prompt: str
    dialogue: str
    reason: str


@dataclass(frozen=True)
class ScoringCriterion:
    """信息量和流畅度页的 prompt/评分标准字段（key-value）。"""

    system: str
    field: str
    dialogue_scope: str
    dialogue_content: str
    diag_info_criterion: str
    diag_flue_criterion: str
    question_surfix: str


@dataclass(frozen=True)
class GoldData:
    cases: list[GoldCase]
    fakes: list[FakeCase]
    criterion: ScoringCriterion


def _to_int_score(v) -> int | None:
    """金标准分转 int；非 0–10 整数返回 None。"""
    if v is None:
        return None
    try:
        f = float(v)
    except (TypeError, ValueError):
        return None
    i = int(round(f))
    if i != f or not (0 <= i <= 10):
        return None
    return i


def read_gold(path: Path | str = GOLD_XLSX) -> GoldData:
    """读取并校验 gold xlsx。评分标准缺关键字段时抛 ValueError（原则 I）。"""
    path = Path(path)
    if not path.exists():
        raise FileNotFoundError(f"gold xlsx 不存在: {path}")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

    criterion = _read_criterion(wb)
    cases = _read_summary(wb)
    fakes = _read_fakeasr(wb)
    wb.close()
    return GoldData(cases=cases, fakes=fakes, criterion=criterion)


def _read_criterion(wb) -> ScoringCriterion:
    ws = wb[SHEET_CRITERION]
    kv: dict[str, str] = {}
    for row in ws.iter_rows(values_only=True):
        if not row or row[0] is None:
            continue
        key = str(row[0]).strip()
        val = "" if len(row) < 2 or row[1] is None else str(row[1])
        kv[key] = val
    missing = [k for k in REQUIRED_CRITERION_KEYS if not kv.get(k, "").strip()]
    if missing:
        raise ValueError(
            f"'{SHEET_CRITERION}' 页缺少必需字段或为空: {missing}（原则 I：禁止用硬编码摘要兜底）"
        )
    return ScoringCriterion(
        system=kv["system"],
        field=kv["field"],
        dialogue_scope=kv["dialogue_scope"],
        dialogue_content=kv["dialogue_content"],
        diag_info_criterion=kv["diag_info_criterion"],
        diag_flue_criterion=kv["diag_flue_criterion"],
        question_surfix=kv["question_surfix"],
    )


def _read_summary(wb) -> list[GoldCase]:
    ws = wb[SHEET_SUMMARY]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError(f"'{SHEET_SUMMARY}' 页为空")
    header = [str(h).strip() if h is not None else "" for h in rows[0]]

    def col(name: str) -> int:
        if name not in header:
            raise ValueError(f"'{SHEET_SUMMARY}' 页缺少列: {name}")
        return header.index(name)

    i_id, i_info, i_flue, i_usage = col("编号"), col("信息量"), col("流畅度"), col("用途")
    i_lang = header.index("语言") if "语言" in header else None

    cases: list[GoldCase] = []
    for r in rows[1:]:
        if not r or r[i_id] is None:
            continue
        case_id = str(r[i_id]).strip()
        info = _to_int_score(r[i_info])
        flue = _to_int_score(r[i_flue])
        usage = str(r[i_usage]).strip() if r[i_usage] is not None else ""
        lang = (
            str(r[i_lang]).strip()
            if i_lang is not None and i_lang < len(r) and r[i_lang] is not None
            else None
        )
        # info/flue 可能缺失（如仅看图说话）；保留 None，由 build_dataset 决定取舍
        cases.append(
            GoldCase(
                case_id=case_id,
                info_gold=info if info is not None else -1,
                flue_gold=flue if flue is not None else -1,
                usage=usage,
                language=lang,
            )
        )
    return cases


def _read_fakeasr(wb) -> list[FakeCase]:
    ws = wb[SHEET_FAKE]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = [str(h).strip() if h is not None else "" for h in rows[0]]

    def idx(name: str) -> int | None:
        return header.index(name) if name in header else None

    i_info, i_flue = idx("信息量"), idx("流畅度")
    i_prompt, i_dlg, i_reason = idx("提示词"), idx("对话"), idx("理由")

    fakes: list[FakeCase] = []
    for r in rows[1:]:
        if not r or i_dlg is None or i_dlg >= len(r) or r[i_dlg] is None:
            continue
        info = _to_int_score(r[i_info]) if i_info is not None else None
        flue = _to_int_score(r[i_flue]) if i_flue is not None else None
        if info is None or flue is None:
            continue
        fakes.append(
            FakeCase(
                info_gold=info,
                flue_gold=flue,
                gen_prompt=str(r[i_prompt]) if i_prompt is not None and r[i_prompt] else "",
                dialogue=str(r[i_dlg]),
                reason=str(r[i_reason]) if i_reason is not None and r[i_reason] else "",
            )
        )
    return fakes
